"""Build the Equity Price Targets workbook from data.json.

Sheets: Dashboard | Price Targets | Tickers
All derived columns are live Excel formulas; raw fetched values are inputs.
"""
import json, datetime, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

DATA = sys.argv[1] if len(sys.argv) > 1 else "data.json"
OUT = sys.argv[2] if len(sys.argv) > 2 else "Equity_Price_Targets.xlsx"

d = json.load(open(DATA))
rows = d["rows"]
# populated rows first (by market cap desc), pending rows after (alphabetical)
pop = sorted([r for r in rows if r.get("price")], key=lambda r: -(r.get("market_cap") or 0))
pend = sorted([r for r in rows if not r.get("price")], key=lambda r: r["ticker"])
rows = pop + pend

ARIAL = "Arial"
F = lambda **kw: Font(name=ARIAL, **kw)
HDR_FILL = PatternFill("solid", fgColor="1F3864")
BAND_FILL = PatternFill("solid", fgColor="F2F2F2")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
THIN = Border(bottom=Side(style="thin", color="D9D9D9"))

wb = Workbook()

# ---------------- Price Targets sheet ----------------
ws = wb.create_sheet("Price Targets")
headers = [
    ("Ticker", 9), ("Company", 24), ("Industry", 26), ("Country", 14),
    ("Market Cap ($B)", 15), ("Price ($)", 11), ("Avg Target ($)", 13),
    ("Upside to Avg (%)", 15), ("Min Target ($)", 13), ("Upside to Min (%)", 15),
    ("Max Target ($)", 13), ("Upside to Max (%)", 15), ("# Analysts", 10),
    ("Annual Div/Share ($)", 17), ("Div Yield (%)", 12), ("Ex-Div Date", 13),
    ("Earnings Date", 14), ("TTM EPS ($)", 11), ("TTM P/E (x)", 11), ("Data As Of", 12),
]
for c, (h, w) in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = F(bold=True, color="FFFFFF", size=10)
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(c)].width = w
ws.row_dimensions[1].height = 30

def dt(s):
    return datetime.datetime.strptime(s, "%Y-%m-%d") if s else None

for i, r in enumerate(rows):
    rw = i + 2
    mc = r.get("market_cap")
    vals = {
        1: r["ticker"], 2: r.get("name"), 3: r.get("industry"), 4: r.get("country"),
        5: round(mc / 1e9, 2) if mc else None,
        6: r.get("price"), 7: r.get("target_avg"),
        8: f'=IF(AND(ISNUMBER(F{rw}),ISNUMBER(G{rw}),F{rw}>0),G{rw}/F{rw}-1,"")',
        9: r.get("target_min"),
        10: f'=IF(AND(ISNUMBER(F{rw}),ISNUMBER(I{rw}),F{rw}>0),I{rw}/F{rw}-1,"")',
        11: r.get("target_max"),
        12: f'=IF(AND(ISNUMBER(F{rw}),ISNUMBER(K{rw}),F{rw}>0),K{rw}/F{rw}-1,"")',
        13: r.get("analysts"), 14: r.get("div_rate"),
        15: f'=IF(AND(ISNUMBER(F{rw}),ISNUMBER(N{rw}),F{rw}>0),N{rw}/F{rw},"")',
        16: dt(r.get("ex_div")), 17: dt(r.get("earnings")), 18: r.get("eps_ttm"),
        19: f'=IF(AND(ISNUMBER(F{rw}),ISNUMBER(R{rw}),R{rw}>0),F{rw}/R{rw},"")',
        20: dt(r.get("as_of")) or (dt(d["as_of"][:10]) if r.get("price") else None),
    }
    for c, v in vals.items():
        cell = ws.cell(row=rw, column=c, value=v)
        cell.font = F(size=10)
        cell.border = THIN
        if i % 2 == 1:
            cell.fill = BAND_FILL
    if not r.get("price"):
        ws.cell(row=rw, column=2, value="awaiting first automated refresh").font = F(size=10, italic=True, color="999999")

n = len(rows) + 1
fmt = {5: '$#,##0.0', 6: '$#,##0.00', 7: '$#,##0.00', 8: '0.0%', 9: '$#,##0.00',
       10: '0.0%', 11: '$#,##0.00', 12: '0.0%', 13: '0', 14: '$0.00', 15: '0.00%',
       16: 'mmm d, yyyy', 17: 'mmm d, yyyy', 18: '$0.00', 19: '0.0"x"', 20: 'mmm d, yyyy'}
for c, f in fmt.items():
    for rw in range(2, n + 1):
        ws.cell(row=rw, column=c).number_format = f

# conditional formatting: upside green / downside red on the three upside columns
for col in ("H", "J", "L"):
    rng = f"{col}2:{col}{n}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"],
        font=Font(name=ARIAL, size=10, color="006100", bold=(col == "H"))))
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"],
        font=Font(name=ARIAL, size=10, color="9C0006", bold=(col == "H"))))
# earnings within next 7 days -> yellow fill
ws.conditional_formatting.add(f"Q2:Q{n}", FormulaRule(
    formula=[f'AND(ISNUMBER(Q2),Q2>=TODAY(),Q2-TODAY()<=7)'], fill=YELLOW))

ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:T{n}"

# ---------------- Dashboard ----------------
db = wb.active
db.title = "Dashboard"
db.sheet_view.showGridLines = False
for col, w in zip("ABCD", (34, 16, 30, 16)):
    db.column_dimensions[col].width = w
t = db["A1"]; t.value = "Equity Price Targets — Dashboard"; t.font = F(bold=True, size=16, color="1F3864")
db["A2"] = "Analyst consensus tracker · data via Yahoo Finance · auto-refreshed each trading day"
db["A2"].font = F(size=10, italic=True, color="666666")

P = "'Price Targets'"
stats = [
    ("Tickers tracked", f'=COUNTA({P}!A2:A{n})', '0'),
    ("Tickers with live data", f'=COUNT({P}!F2:F{n})', '0'),
    ("Median upside to avg target", f'=MEDIAN({P}!H2:H{n})', '0.0%'),
    ("Average dividend yield (payers)", f'=AVERAGEIF({P}!O2:O{n},">0")', '0.00%'),
    ("Companies reporting in next 7 days", f'=COUNTIFS({P}!Q2:Q{n},">="&TODAY(),{P}!Q2:Q{n},"<="&TODAY()+7)', '0'),
    ("Largest upside to avg target", f'=MAX({P}!H2:H{n})', '0.0%'),
    ("→ ticker", f'=INDEX({P}!A2:A{n},MATCH(MAX({P}!H2:H{n}),{P}!H2:H{n},0))', 'General'),
    ("Smallest upside / largest downside", f'=MIN({P}!H2:H{n})', '0.0%'),
    ("→ ticker", f'=INDEX({P}!A2:A{n},MATCH(MIN({P}!H2:H{n}),{P}!H2:H{n},0))', 'General'),
    ("Combined market cap ($T)", f'=SUM({P}!E2:E{n})/1000', '$#,##0.00'),
]
for i, (label, formula, nf) in enumerate(stats):
    rw = 4 + i
    db.cell(row=rw, column=1, value=label).font = F(size=11)
    c = db.cell(row=rw, column=2, value=formula)
    c.font = F(size=11, bold=True); c.number_format = nf
db.cell(row=15, column=1, value=f"Data as of: {d['as_of']}").font = F(size=9, color="999999")
db.cell(row=16, column=1, value="Legend: green = trading below analyst targets (upside), red = above (downside). "
        "Yellow earnings dates are within 7 days.").font = F(size=9, color="999999")

# ---------------- Tickers config ----------------
tk = wb.create_sheet("Tickers")
tk.column_dimensions["A"].width = 12; tk.column_dimensions["B"].width = 18; tk.column_dimensions["C"].width = 70
tk["A1"] = "Ticker"; tk["B1"] = "Group"
for c in ("A1", "B1"):
    tk[c].font = F(bold=True, color="FFFFFF", size=10); tk[c].fill = HDR_FILL
tk["C1"] = "HOW TO ADD A STOCK"
tk["C1"].font = F(bold=True, size=10)
notes = [
    "This tab mirrors the ticker universe used by the automated refresh.",
    "To add or remove a stock: edit the tickers.txt file in the GitHub repo (one symbol per line),",
    "or simply tell Claude (e.g. “add NFLX and COIN to my price targets sheet”).",
    "The next scheduled refresh picks the change up automatically — no other edits needed.",
    "Note: use Yahoo-style symbols (BRK-B, not BRK.B) when editing tickers.txt.",
]
for i, s in enumerate(notes):
    tk.cell(row=2 + i, column=3, value=s).font = F(size=9, color="444444")
groups = {r["ticker"]: ("International (ADR)" if r["ticker"].replace(".", "-") in
          {"TSM","ASML","NVO","NSRGY","SAP","TM","SONY","AZN","NVS","SHEL","HSBC","UL",
           "BABA","TCEHY","LVMUY","RHHBY","TTE","BHP"} else "S&P 100 / US") for r in rows}
for i, r in enumerate(sorted(rows, key=lambda x: x["ticker"])):
    tk.cell(row=2 + i, column=1, value=r["ticker"]).font = F(size=10)
    tk.cell(row=2 + i, column=2, value=groups[r["ticker"]]).font = F(size=10, color="666666")
tk.freeze_panes = "A2"

wb.move_sheet("Price Targets", offset=-1)  # Dashboard, Price Targets, Tickers -> ensure order
wb._sheets = [wb["Dashboard"], wb["Price Targets"], wb["Tickers"]]
wb.save(OUT)
print(f"saved {OUT}: {len(rows)} rows ({len(pop)} populated)")
